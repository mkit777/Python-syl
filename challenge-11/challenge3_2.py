#!-*-coding:utf-8-*-
#!/usr/bin/env python3

from openpyxl import load_workbook,Workbook
FILE = 'D:/workspace/python/courses.xlsx'
def combine():
    wb = load_workbook(FILE)
    s_stds = wb['students']
    s_times = wb['time']
    s_combine = wb.create_sheet('combine')
    for row in range(1,s_stds.max_row+1):

        #将前三行复制过去
        for col in range(1,s_stds.max_column+1):
            val_src = s_stds.cell(row,col).value
            s_combine.cell(row,col,val_src)

        #在s_time寻找相等的值
        for data_time in s_times.values:
            if  s_combine.cell(row,col-1).value == data_time[1]:
                s_combine.cell(row,col+1,data_time[2])
                break
        else:
            s_combine.cell(row,col+1,data_time[2])

    wb.save(FILE)    
    wb.close()       

def split():
    wb = load_workbook(FILE)
    st = wb['combine']
    head = list(get_values(st[1]))
    wbs={}
    for i,row in enumerate(st.values):
        if i == 0:
            continue
        year = row[0].year
        if year not in wbs:
            nwb = Workbook()
            wbs[year] = nwb
            sheet = nwb.create_sheet()
            sheet.append(head)
            nwb.save('D:/'+str(year)+'.xlsx')
            nwb.close()
        wb = wbs.get(year)
        sheet = wb.active
        sheet.append(row)

    print(sheet.values)
    for k,v in wbs.items():
        print(k,v)
        v.save('D:/'+str(k)+'.xlsx')
        v.close()

def get_values(row):
    for cell  in row:
        yield cell.value


if __name__ == '__main__':
    #combine()
    split()